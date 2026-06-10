import pandas as pd
import customtkinter
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from buscar_atras import buscar_datos_nave_hacia_atras

def obtener_ruta_salida(base_path):
    """Genera el nombre Formato_Tom(n).xlsx si el archivo ya existe."""
    filename = "Formato_Tom.xlsx"
    full_path = os.path.join(base_path, filename)
    counter = 1
    
    while os.path.exists(full_path):
        full_path = os.path.join(base_path, f"Formato_Tom({counter}).xlsx")
        counter += 1
    return full_path




def procesar_datos():
    ruta_input = entry_archivo.get()
    fecha_ini_str = entry_inicio.get()
    fecha_ter_str = entry_fin.get()
    ruta_output_dir = entry_destino.get()

    # Formato Fecha (dd/mm/yyyy h:mm)
    # Nota: En Excel 's' es para segundos, usa 'mm' para minutos
    formato_fecha = 'dd/mm/yyyy h:mm'

    # Formato decimal (Numérico, 2 decimales, sin miles)
    formato_dec = '0.00'

    # Formato Entero (Entero, con separador de miles)
    # El '#' indica dígito opcional, el '0' indica dígito obligatorio
    formato_entero = '#,##0'
    formato_multa = '#,##0.00'

    

    border_horizontal = Border(
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"))

    border_completo = Border(
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000")
)
    
    

    relleno_rosa =   PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    relleno_celeste = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")

    relleno_gris =    PatternFill(start_color="F2F2F2", end_color = "F2F2F2", fill_type="solid") 

    #Color Amarillo Claro (Hexadecimal: FFFFE0 o FFF9C4 para un crema suave)
    relleno_amarillo = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")                                                                                  

    # Estilo Liviano
    fuente_liviana = Font(name='Aptos Narrow', size=9, color='808080')
    alineacion_sangria = Alignment(indent=1, vertical='center')


    fuente_resultados = Font(name='Aptos NArrow', size=11, bold=True)
    fuente_datos = Font(name = "Aptos Narrow", size = 10, bold = False)

    # 2. Definición de Borde Nulo (para asegurar que el cuerpo no tenga nada)
    sin_bordes = Border(
    top=Side(border_style=None),
    bottom=Side(border_style=None),
    left=Side(border_style=None),
    right=Side(border_style=None))

    if not all([ruta_input, fecha_ini_str, fecha_ter_str, ruta_output_dir]):
        messagebox.showwarning("Error", "Por favor complete todos los campos.")
        return

    try:
        # Convertir fechas de entrada
        f_inicio = datetime.strptime(fecha_ini_str, "%d/%m/%Y")
        f_termino = datetime.strptime(fecha_ter_str, "%d/%m/%Y")

        # Cargar Excel (Asumiendo que no tiene encabezados o detectándolos)
        df = pd.read_excel(ruta_input)
        
        # Mapeo de columnas (A=0, C=2, L=11, M=12, N=13)
        # Col C: Nave, Col L: TRG, Col M: Atraque
        col_nave = df.columns[2]
        col_servicio = df.columns[4]
        col_tipo_nave = df.columns[5]
        col_term = df.columns[6]
        col_sitio = df.columns[7]
        col_loa = df.columns[9]
        col_trg = df.columns[11]
        col_atraque = df.columns[12]
        col_desatraque = df.columns[13]
        col_ton = df.columns[31]
        col_xs = df.columns[53]
        col_cu_met = df.columns[56]
        col_zn = df.columns[57]
        col_cu = df.columns[59]
        col_cs = df.columns[61]
        col_break = df.columns[64]
        
        # Convertir columna de atraque a datetime
        df[col_atraque] = pd.to_datetime(df[col_atraque],  dayfirst=True,   # Indica que el día va primero (dd/mm/aaaa)
        errors='coerce'  # Si hay un error (texto o celda vacía), lo convierte en "NaT" (Not a Time) en lugar de fallar
        )
        
        # Filtrar por rango,  que el terminal contenga "ATI"
        mask = (
            (df[col_atraque] >= f_inicio) & 
            (df[col_atraque] <= f_termino) & 
            (df[col_term].astype(str).str.contains("ATI", na=False)) # Nuevo filtro ATI
        )
       
        df_filtrado = df.loc[mask].sort_values(by=col_atraque)

        if df_filtrado.empty:
            messagebox.showinfo("Sin resultados", "No se encontraron registros en el rango especificado")
            return

        # Crear nuevo libro
        wb = Workbook()
        # Eliminar hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Crear hoja resumen al principio del libro
        ws_res = wb.create_sheet(title="Res", index=0)
        ws_res.sheet_view.showGridLines = False

        hojas_naves = []
        
         # Encabezados hoja Res desde B1 hasta R1
        encabezados_res = [
            "Nº",
            "Nave",
            "Tipo Nave",
            "TRG",
            "Sitio",
            "Primera Espia",
            "Ultima Espia",
            "T.O",
            "B. Dom",
            "F.Rend",
            "TOM",
            "TNT",
            "GAP",
            "Multa Inicial",
            "Ton",
            "Vel (ton/hr)",
            "TIPO DE CARGA"
        ]
        ws_res.column_dimensions['B'].width = 4
        ws_res.column_dimensions['C'].width = 24
        ws_res.column_dimensions['D'].width = 16
        ws_res.column_dimensions['E'].width = 10
        ws_res.column_dimensions['F'].width = 7
        ws_res.column_dimensions['G'].width = 16
        ws_res.column_dimensions['H'].width = 16
        ws_res.column_dimensions['I'].width = 8
        ws_res.column_dimensions['J'].width = 9
        ws_res.column_dimensions['K'].width = 9
        ws_res.column_dimensions['L'].width = 8
        ws_res.column_dimensions['M'].width = 8
        ws_res.column_dimensions['N'].width = 8
        ws_res.column_dimensions['O'].width = 12
        ws_res.column_dimensions['P'].width = 10
        ws_res.column_dimensions['Q'].width = 10
        ws_res.column_dimensions['R'].width = 35

        
        for i, texto in enumerate(encabezados_res, start=2):  # B=2 hasta R=18
            celda = ws_res.cell(row=1, column=i, value=texto)
            celda.font = Font(name="Aptos Narrow", size=10, bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
            celda.border = border_completo
        

        for index, row in df_filtrado.iterrows():
            nombre_nave = str(row[col_nave]).strip()

            # Revisar si el TRG actual viene vacío
            trg_val = pd.to_numeric(row[col_trg], errors='coerce')
            loa_val = pd.to_numeric(row[col_loa], errors='coerce')

            datos_recuperados = None
            es_segunda_recalada_o_repetida = False

            if pd.isna(trg_val):
                datos_recuperados = buscar_datos_nave_hacia_atras(
                    df=df,
                    index_actual=index,
                    col_nave=col_nave,
                    col_trg=col_trg,
                    col_loa=col_loa,
                    max_filas=20
                )

                if datos_recuperados is not None:
                    trg_val = datos_recuperados["trg"]
                    loa_val = datos_recuperados["loa"]
                    es_segunda_recalada_o_repetida = True

            # Crear nombre base de hoja
            sheet_name_base = "".join(x for x in nombre_nave if x.isalnum() or x in " -_")[:31]

            # Si se recuperó TRG desde una fila anterior, agregar (2), (3), etc.
            if es_segunda_recalada_o_repetida:
                contador = 2

                while True:
                    sufijo = f"({contador})"
                    largo_max_base = 31 - len(sufijo)
                    sheet_name = f"{sheet_name_base[:largo_max_base]}{sufijo}"

                    if sheet_name not in wb.sheetnames:
                        break

                    contador += 1
            else:
                sheet_name = sheet_name_base

                # Si por cualquier razón ya existe una hoja con ese nombre,
                # también se numera para evitar error de Excel.
                if sheet_name in wb.sheetnames:
                    contador = 2

                    while True:
                        sufijo = f"({contador})"
                        largo_max_base = 31 - len(sufijo)
                        sheet_name = f"{sheet_name_base[:largo_max_base]}{sufijo}"

                        if sheet_name not in wb.sheetnames:
                            break

                        contador += 1

            ws = wb.create_sheet(title=sheet_name)
            hojas_naves.append(sheet_name)

            # 1. Desactivar cuadrícula
            ws.sheet_view.showGridLines = False

            #aca deseo que desde la columna H hasta la columna G tengan un ancho de 10
            ws.column_dimensions['B'].width = 27
            ws.column_dimensions['C'].width = 21

            # Ajustar ancho de 10 para columnas desde la H hasta la O
            columnas_cuadro = ['G','H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
            for letra in columnas_cuadro:
                ws.column_dimensions[letra].width = 10

            # --- Formato Numérico para el Cuadro de Bodegas (H20:O27) ---
            # Recorremos desde la fila 20 hasta la 27
            for r in range(20, 28): 
                # Recorremos desde la columna 8 (H) hasta la 15 (O)
                for c in range(7, 15): 
                    celda_cuadro = ws.cell(row=r, column=c)
                    
                    # Aplicamos el formato definido previamente
                    celda_cuadro.number_format = formato_multa
                    
                    # Opcional: Alinear a la derecha para que los decimales queden ordenados
                    celda_cuadro.alignment = Alignment(horizontal='right')

            
            # --- DATOS GENERALES (Texto) ---
            ws.cell(row=3, column=2, value="Nave")
            ws.cell(row=3, column=3, value=nombre_nave)

            ws.cell(row=4, column=2, value="Servicio")
            ws.cell(row=4, column=3, value=str(row[col_servicio]))

            ws.cell(row=5, column=2, value="Viaje")
            #ws.cell(row=5, column=3, value=row[col_viaje]) # Asumiendo columna de viaje

            ws.cell(row=17, column=4, value="Ton/hora")
            ws.cell(row=18, column=4, value="Ton")
            ws.cell(row=19, column=4, value="Hr")
            ws.cell(row=20, column=4, value="Hr")      

            c_loa = ws.cell(row=7, column=3, value=float(row[col_loa]))
            c_loa.number_format = formato_dec
            # LOA
            ws.cell(row=7, column=2, value="LOA")

            if pd.isna(loa_val):
                c_loa = ws.cell(row=7, column=3, value="")
            else:
                c_loa = ws.cell(row=7, column=3, value=float(loa_val))
                c_loa.number_format = formato_dec

            # TRG
            ws.cell(row=8, column=2, value="TRG")

            if pd.isna(trg_val):
                c_trg = ws.cell(row=8, column=3, value="")
            else:
                c_trg = ws.cell(row=8, column=3, value=int(trg_val))
                c_trg.number_format = formato_entero

            ws.cell(row=9, column=2, value="Sitio")        
            ws.cell(row=9, column=3, value=int(row[col_sitio]))

            ws.cell(row=10, column=2, value="Fecha Requerida de Atraque")
            c_fech_req = ws.cell(row=10, column=3)
            c_fech_req.number_format = formato_fecha

            ws.cell(row=11, column=2, value="Fecha amarra primera espia")        
            c_atraque = ws.cell(row=11, column=3, value=row[col_atraque])
            c_atraque.number_format = formato_fecha
            
            ws.cell(row=12, column=2, value="Fecha desamarra ultima espia")        
            c_desatraque = ws.cell(row=12, column=3, value=row[col_desatraque])
            c_desatraque.number_format = formato_fecha

            ws.cell(row=13, column=2, value="Tiempo de Espera")
            ws.cell(row=13, column=3, value='=IF(ISBLANK(C10),"",24*(C11-C10))')

            ws.cell(row=14, column=2, value="Tiempo de ocupación")
            tiempo_ocupacion = ws.cell(row=14, column=3, value="=24*(C12-C11)")
            tiempo_ocupacion.number_format = formato_dec

            for celda in ws["B14:C14"][0]: 
                celda.fill = relleno_rosa

            ws.cell(row=16, column=2, value="Toneladas Transferidas")
            ton_val = pd.to_numeric(row[col_ton], errors="coerce")
            ton_val = float(0 if pd.isna(ton_val) else ton_val)

            tons = ws.cell(row=16, column=3, value=ton_val)
            tons.number_format = formato_entero
            
            ws.cell(row=17, column=2, value="Factor de Rendimiento")
            
            ws.cell(row=18, column=2, value="Bodega Dominante")
            ton_dominante = ws.cell(row=18, column=3, value="=MAX(G27:M27)")
            ton_dominante.number_format = formato_dec

            ws.cell(row=27, column=6, value="Totales")
           
           # --- DATOS GENERALES (Fila 6: Tipo Nave) ---
            ws.cell(row=6, column=2, value="Tipo Nave")
            
            # Paso 1: Convertir a número de forma segura (evita el error de base 10)
            tipo_nave_val = pd.to_numeric(row[col_tipo_nave], errors='coerce')
            #deseo que ceniza_soda sea numero decimal ¿como lo hago?
            tipo_nave_val = int(0 if pd.isna(tipo_nave_val) else tipo_nave_val) # Si es vacío, asume 0

            ceniza_soda = pd.to_numeric(row[col_cs], errors='coerce')
            ceniza_soda = float(0 if pd.isna(ceniza_soda) else ceniza_soda)
            if ceniza_soda>0:
                ws.cell(row=16,column=7, value = "CENIZA DE SODA")
                ws.cell(row=17, column=3, value=96)

            #porque siempre cobre_met=0, cuando en ocasiones no lo es
            #cobre metalico
            
            cobre_met = pd.to_numeric(row[col_cu_met], errors="coerce")
            cobre_met = int(0 if pd.isna(cobre_met) else cobre_met)
            print ("COBRE =", cobre_met, "toneladas", ton_val)
            
            #break bulk
            break_bulk = pd.to_numeric(row[col_break], errors="coerce")
            break_bulk = float(0 if pd.isna(break_bulk) else break_bulk)
            
            if break_bulk > 0 and abs(break_bulk - ton_val) < 0.01 and cobre_met==0:
                ws.cell(row=16, column=7, value="BREAK BULK")
                ws.cell(row=17, column=3, value=40)

            if break_bulk > 0 and abs(break_bulk - cobre_met) < 0.01:
                ws.cell(row=16, column=7, value="COBRE METALICO")
                ws.cell(row=17, column=3, value=120)

            conc_cu = pd.to_numeric(row[col_cu], errors="coerce")
            conc_cu = int(0 if pd.isna(conc_cu) else conc_cu)
            if conc_cu > 0 and conc_cu == ton_val:
                ws.cell(row=16,column=7, value = "CONCENTRADO DE COBRE")
                ws.cell(row=17, column=3, value=250)

            conc_zn = pd.to_numeric(row[col_zn], errors="coerce")
            conc_zn = int(0 if pd.isna(conc_zn) else conc_zn)
            if conc_zn > 0 and conc_zn == ton_val:
                ws.cell(row=16,column=7, value = "CONCENtRADO DE ZINC")
                ws.cell(row=17, column=3, value=250)

            if conc_zn > 0 and conc_cu >0:
                ws.cell(row=16,column=7, value = "CONC. ZINC + CONC. CU")
                ws.cell(row=17, column=3, value=250)

            # Caso Tipo Nave 32 (Container Ship)
            if tipo_nave_val == 32:
                ws.cell(row=6, column=3, value="Container Ship")
                ws.cell(row=16, column = 7, value = "CONTENEDORES")
                # Sumamos columnas 32 a 39 (el límite 40 es para incluir la 39)
                suma_embarque = pd.to_numeric(row.iloc[32:40], errors='coerce').fillna(0).sum()
                ws.cell(row=17, column=3, value=21)
                ws.cell(row=17, column=4, value="ctnr/hora")
                ws.cell(row=18, column=4, value="ctnr")

                teus=ws.cell(row=18, column=3, value=suma_embarque)
                teus.number_format = formato_entero

            # Caso Tipo Nave 13 (Bulk Carrier)
            elif tipo_nave_val == 13:
                ws.cell(row=6, column=3, value="Bulk Carrier")
                
                # Conversión segura para Zinc
                val_zn = pd.to_numeric(row[col_zn], errors='coerce')
                val_zn = 0 if pd.isna(val_zn) else val_zn
                
                if val_zn > 0:
                    ws.cell(row=17, column=3, value=250)

            elif tipo_nave_val == 11:
                ws.cell(row=6, column=3, value="General Cargo Ship")

            ws.cell(row=19, column=2, value="Tiempo de Ocupacion Maximo")
            tom = ws.cell(row=19, column=3, value="=IFERROR(C18/C17, 0)+2")
            tom.number_format = formato_dec
            

            ws.cell(row=20, column=2, value="Tiempo de Inactividad")
            tnt = ws.cell(row=20, column=3, value="=SUM(C21:C27)") 
            tnt.number_format = formato_dec

            ws.cell(row=21, column=2, value="Atraque y Recepcion Nave")
            ws.cell(row=22, column=2, value="Colacion")
            ws.cell(row=23, column=2, value="Descanso Legal Amanecida")
            ws.cell(row=24, column=2, value="Tiempo Muerto Remate")
            ws.cell(row=25, column=2, value="Espera de carga / camiones")
            ws.cell(row=25, column=2, value="Draft")
            ws.cell(row=26, column=2, value="Inspeccion de Bodegas")
            ws.cell(row=27, column=2, value="")
            ws.cell(row=28, column=2, value="")
            ws.cell(row=29, column=2, value="")
            ws.cell(row=30, column=2, value="")
            ws.cell(row=31, column=2, value="")
            
            # Aplicar formato liviano + fondo amarillo al rango deseado
            # Cambia "B18:B31" por el rango que necesites resaltar
            for fila in ws["B21:C31"]:
                for celda in fila:
                    celda.font = fuente_liviana
                    celda.alignment = alineacion_sangria
                    celda.fill = relleno_amarillo
            for fila in ws["C21:C31"]:
                for celda in fila:
                    celda.number_format = formato_dec

            ws.cell(row=32, column=2, value="Tiempo Maximo Permitido (Real)")
            t_max= ws.cell(row=32, column=3, value="=C19+C20")
            t_max.number_format = '0.00'

            ws.cell(row=33, column=2, value="GAP")
            gap = ws.cell(row=33, column=3, value='=IF(C17=0,"Sin determinar aun",C32-C14)')
            gap.number_format = '0.00'

            ws.cell(row=34, column=2, value="*Si GAP (+) no aplica multa. Si GAP (-) aplica multa")

            for celda in ws["B34:C34"][0]: 
                celda.fill = relleno_celeste

            tit_tom= ws.cell(row=36, column=2, value="Multa Tom")
            tit_tom.font = fuente_resultados

            ws.cell(row=37, column=2, value="Multa Base")
            ws.cell(row=37, column=3, value=0.035)

            ws.cell(row=38, column=2, value="Factor USPPI")
            ws.cell(row=38, column=3, value=1.803)

            ws.cell(row=39, column=2, value="Multa Final")
            m_final= ws.cell(row=39, column=3, value="=C37*C38")
            m_final.number_format = '0.000'
        
            calc_tom= ws.cell(row=41, column=2, value="Calculo Multa Tom")
            calc_tom.font = fuente_resultados

            ws.cell(row=42, column=2, value="Multa Final")
            ws.cell(row=42, column=3, value="=C39")

            ws.cell(row=43, column=2, value="Horas Exceso")
            ws.cell(row=43, column=3, value='=IFERROR(C33*-1, "Sin definir aun")')
            ws.cell(row=44, column=2, value="TRG")
            trg=ws.cell(row=44, column=3, value="=C8")
            trg.number_format = formato_entero

            ws.cell(row=45, column=2, value="MULTA TOM").border = border_horizontal
            m_apl = ws.cell(row=45, column=3, value='=IF(C20=0,"Sin Revision",IF(C43<0,0,C44*C43*C42))')
           
            m_apl.number_format = formato_multa
            m_apl.border = border_horizontal

            ws.cell(row=47, column=2, value="RENDIMIENTO").border = border_horizontal
            rend = ws.cell(row=47, column= 3, value = "=C16/(C14-C20)")
            rend.number_format = formato_multa
            rend.border = border_horizontal
        
            for celda in ws[45][1:3]: # Accede directamente a la fila 45, columnas B(1) y C(2)
                celda.fill = relleno_rosa
            
            for celda in ws["B47:C47"][0]: 
                celda.fill = relleno_celeste

        # --- CONFIGURACIÓN DEL CUADRO BODEGAS  (F18:N27) ---
            
            # 1. Definir los encabezados del cuadro (Fila 19, de F a N)
            encabezados_cuadro = ["", "BOD 1", "BOD 2", "BOD 3", "BOD 4", "BOD 5", "BOD 6", "BOD 7", "TOTALES"]
            

            for i, texto in enumerate(encabezados_cuadro):
                celda = ws.cell(row=19, column=6 + i, value=texto)
                celda.font = fuente_resultados
                celda.alignment = Alignment(horizontal='right')
                celda.border = border_horizontal

            # 4. Configuración del Cuerpo y Totales (Filas 19 a 27)
            for r in range(20, 29): # De 20 hasta 28
                for c in range(6, 15): # Columnas F a N
                    celda_actual = ws.cell(row=r, column=c)
                    
                    if r == 27:
                        # Fila de TOTALES: Aplicar borde horizontal arriba y abajo
                        celda_actual.font = fuente_resultados
                        celda_actual.border = border_horizontal
                    else:
                        # Cuerpo de datos: Sin bordes, solo fuente normal
                        celda_actual.font = fuente_datos
                        celda_actual.border = sin_bordes

            #Revisar al parecer hay un error porque al abri excel envia un mensaje de error para reparar formula

            # 3. FÓRMULAS DE SUMA POR FILA (Columna N)
            # Sumar desde la columna F (6) hasta la M (13) para cada fila
            for r in range(19, 27): # Filas de datos (debajo del encabezado y sobre el total)
                # Ejemplo: =SUM(F19:M19)
                ws.cell(row=r, column=14).value = f"=SUM(G{r}:M{r})"

            # 4. FÓRMULAS DE SUMA POR COLUMNA (Fila 27)
            # Sumar verticalmente desde la fila 18 a la 26 para cada columna
            columnas_a_sumar = ['G', 'H', 'I', 'J', 'K', 'L', 'M','N']

            for letra in columnas_a_sumar:
                # La 'G' es el índice 0 en nuestra lista. 
                # Para que caiga en la columna 7 de Excel: 0 + 7 = 7.
                col_num = columnas_a_sumar.index(letra) + 7                
                # La suma debe ir desde la fila 20 (primer dato) hasta la 26 (último dato)
                ws.cell(row=27, column=col_num).value =f"=SUM({letra}20:{letra}26)"

        fila_res = 2

        for nombre_hoja in hojas_naves:
            ref_hoja = f"'{nombre_hoja}'"

            ws_res.cell(row=fila_res, column=2, value=fila_res-1)
            ws_res.cell(row=fila_res, column=3, value=f"={ref_hoja}!C3") #Nombre Nave
            ws_res.cell(row=fila_res, column=4, value=f"={ref_hoja}!C6") #Tipo Nave
            
            trg = ws_res.cell(row=fila_res, column=5, value=f"={ref_hoja}!C8") #TRG
            trg.number_format = formato_entero
            trg.alignment = Alignment(horizontal="center", vertical="center")

            sitio= ws_res.cell(row=fila_res, column=6, value=f"={ref_hoja}!C9") #Sitio
            sitio.alignment = Alignment(horizontal="center", vertical="center")

            p_espia= ws_res.cell(row=fila_res, column=7, value=f"={ref_hoja}!C11") #Primera Espia
            p_espia.number_format = formato_fecha
            u_espia= ws_res.cell(row=fila_res, column=8, value=f"={ref_hoja}!C12") #Ultima Espia
            u_espia.number_format = formato_fecha
            # REVISAR
            t_oc= ws_res.cell(row=fila_res, column=9, value=f"={ref_hoja}!C14") #T.O.
            t_oc.number_format = formato_dec
            t_oc.alignment = Alignment(horizontal="center", vertical="center")

            b_dom= ws_res.cell(row=fila_res, column=10, value=f"={ref_hoja}!C18") #B.Dom
            b_dom.number_format = formato_dec
            b_dom.alignment = Alignment(horizontal="center", vertical="center")

            f_rend = ws_res.cell(row=fila_res, column=11, value=f"={ref_hoja}!C17") # F. Rend
            f_rend.number_format = formato_dec
            f_rend.alignment = Alignment(horizontal="center", vertical="center")

            tom= ws_res.cell(row=fila_res, column=12, value=f"={ref_hoja}!C19") #TOM
            tom.number_format = formato_dec
            tom.alignment = Alignment(horizontal="center", vertical="center")

            tnt= ws_res.cell(row=fila_res, column=13, value=f"={ref_hoja}!C20") #TNT
            tnt.number_format = formato_dec
            tnt.alignment = Alignment(horizontal="center", vertical="center")


            gap= ws_res.cell(row=fila_res, column=14, value=f"={ref_hoja}!C43") #GAP
            gap.number_format = formato_dec
            gap.alignment = Alignment(horizontal="center", vertical="center")

            multa= ws_res.cell(row=fila_res, column=15, value=f"={ref_hoja}!C45") #"Multa Inicial"
            multa.number_format = formato_dec
            multa.alignment = Alignment(horizontal="center", vertical="center")

            ton= ws_res.cell(row=fila_res, column=16, value=f"={ref_hoja}!C16") # Ton
            ton.number_format = formato_entero
            ton.alignment = Alignment(horizontal="center", vertical="center")

            vel = ws_res.cell(row=fila_res, column=17, value=f"={ref_hoja}!C47") # Vel
            vel.number_format = formato_dec
            vel.alignment = Alignment(horizontal="center", vertical="center")

            ws_res.cell(row=fila_res, column=18, value=f"={ref_hoja}!G16") #TIPO CARGA

            fila_res += 1
        ultima_fila_res = ws_res.max_row

        # Fila donde irá el total
        fila_total = ultima_fila_res + 2

        ws_res.cell(row=fila_total, column=4, value="TOTALES")  # Columna D
        ws_res.cell(row=fila_total, column=4).font = Font(name="Aptos Narrow", size=12, bold=True)
        ws_res.cell(row=fila_total, column=4).alignment = Alignment(horizontal="right", vertical="center")

        
        for celda in ws_res[fila_total][3:16]: # Accede directamente a la fila 47, columnas D(1) y P(2)
                celda.fill = relleno_gris
                celda.border = border_horizontal
                
        ws_res.cell(row=fila_total, column=4).border = Border(left=Side(border_style="thin", color="000000"),
                                                              top=Side(border_style="thin", color="000000"),
                                                              bottom=Side(border_style="thin", color="000000"))

        celda_total_multa = ws_res.cell(row=fila_total,column=15,value=f"=SUM(O2:O{ultima_fila_res})")
        celda_total_multa.border = border_completo
        celda_total_multa.number_format = formato_multa
        celda_total_multa.font = Font(name="Aptos Narrow", size=12, bold=True)
        celda_total_multa.alignment = Alignment(horizontal="center", vertical="center")

        celda_total_ton = ws_res.cell(row=fila_total, column=16, value=f"=SUM(P2:P{ultima_fila_res})")
        celda_total_ton.border = border_completo
        celda_total_ton.number_format = formato_entero
        celda_total_ton.font = Font(name="Aptos Narrow", size=12, bold=True)
        celda_total_ton.alignment = Alignment(horizontal="center", vertical="center")

        
        for fila in ws_res.iter_rows(
            min_row=2,
            max_row=ultima_fila_res,
            min_col=2,   # B
            max_col=18   # R
        ):
            for celda in fila:
                celda.border = border_completo

        # Guardar archivo
        ruta_final = obtener_ruta_salida(ruta_output_dir)
        wb.save(ruta_final)
        messagebox.showinfo("Éxito", f"Archivo creado en:\n{ruta_final}")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")

# --- Interfaz Gráfica ---
root = tk.Tk()
root.title("Generador de Formato TOM")
root.geometry("500x350")

def seleccionar_archivo():
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    entry_archivo.delete(0, tk.END)
    entry_archivo.insert(0, path)

def seleccionar_destino():
    path = filedialog.askdirectory()
    entry_destino.delete(0, tk.END)
    entry_destino.insert(0, path)

# Layout
tk.Label(root, text="Libro de Origen:").pack(pady=5)
entry_archivo = tk.Entry(root, width=50)
entry_archivo.pack()
tk.Button(root, text="Buscar", command=seleccionar_archivo).pack()

tk.Label(root, text="Fecha Inicio (dd/mm/aaaa):").pack(pady=5)
entry_inicio = tk.Entry(root)
entry_inicio.pack()

tk.Label(root, text="Fecha Término (dd/mm/aaaa):").pack(pady=5)
entry_fin = tk.Entry(root)
entry_fin.pack()

tk.Label(root, text="Carpeta de Destino:").pack(pady=5)
entry_destino = tk.Entry(root, width=50)
entry_destino.pack()
tk.Button(root, text="Buscar", command=seleccionar_destino).pack()

tk.Button(root, text="PROCESAR", command=procesar_datos, bg="green", fg="white", height=2).pack(pady=20)


if __name__ == "__main__":
    root.mainloop()