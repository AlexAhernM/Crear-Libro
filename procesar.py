from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkinter import messagebox
import pandas as pd
from datetime import datetime
from buscar_atras import buscar_datos_nave_hacia_atras
from ruta_salida import obtener_ruta_salida
from encabezados import (formato_fecha, formato_dec, formato_entero, 
                         border_horizontal, border_completo, sin_bordes,  
                         relleno_rosa, relleno_celeste, relleno_gris, relleno_amarillo, 
                         fuente_liviana, alineacion_sangria, fuente_resultados, fuente_datos)
from encabezados import crear_encabezados
from excel_sheets import crear_hojas
def procesar_datos(entry_archivo, entry_destino, entry_fin, entry_inicio):
    ruta_input = entry_archivo.get()
    fecha_ini_str = entry_inicio.get()
    fecha_ter_str = entry_fin.get()
    ruta_output_dir = entry_destino.get()

    if not all([ruta_input, fecha_ini_str, fecha_ter_str, ruta_output_dir]):
        messagebox.showwarning("Error", "Por favor complete todos los campos.")
        return

    try:
        # Convertir fechas de entrada
        f_inicio = datetime.strptime(fecha_ini_str, "%d/%m/%Y")
        f_termino = datetime.strptime(fecha_ter_str, "%d/%m/%Y")

        # Cargar Excel (Asumiendo que no tiene encabezados o detectándolos)
        df = pd.read_excel(ruta_input)
        
        # Mapeo de columnas (A=0, C=2, ...etc)
        col_nave = df.columns[2]
        col_servicio = df.columns[4]
        col_tipo_nave = df.columns[5]
        col_term = df.columns[6]
        col_sitio = df.columns[7]
        col_loa = df.columns[9]
        col_trg = df.columns[11]
        col_atraque = df.columns[12]
        col_desatraque = df.columns[13]
        col_ton = df.columns[31]
        col_xs = df.columns[53]
        col_cu_met = df.columns[56]
        col_zn = df.columns[57]
        col_cu = df.columns[59]
        col_cs = df.columns[61]
        col_break = df.columns[64]
        
        # Convertir columna de atraque a datetime
        df[col_atraque] = pd.to_datetime(df[col_atraque],  dayfirst=True,   # Indica que el día va primero (dd/mm/aaaa)
        errors='coerce'  # Si hay un error (texto o celda vacía), lo convierte en "NaT" (Not a Time) en lugar de fallar
        )
        
        # Filtrar por rango,  que el terminal contenga "ATI"
        mask = (
            (df[col_atraque] >= f_inicio) & 
            (df[col_atraque] <= f_termino) & 
            (df[col_term].astype(str).str.contains("ATI", na=False)) # Nuevo filtro ATI
        )
       
        df_filtrado = df.loc[mask].sort_values(by=col_atraque)

        if df_filtrado.empty:
            messagebox.showinfo("Sin resultados", "No se encontraron registros en el rango especificado")
            return

        # Crear nuevo libro
        wb = Workbook()
        # Eliminar hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Crear hoja resumen al principio del libro
        ws_res = wb.create_sheet(title="Res", index=0)
        ws_res.sheet_view.showGridLines = False

                
        crear_encabezados (ws_res, border_completo)
        

        hojas_naves = crear_hojas(wb, df_filtrado, df,  col_nave, col_trg, col_loa, col_servicio, col_sitio, col_atraque, col_desatraque,
                col_tipo_nave, col_ton, col_cs, col_cu_met, col_break, col_zn, col_cu)

        

        fila_res = 2

        for nombre_hoja in hojas_naves:
            ref_hoja = f"'{nombre_hoja}'"

            ws_res.cell(row=fila_res, column=2, value=fila_res-1)
            ws_res.cell(row=fila_res, column=3, value=f"={ref_hoja}!C3") #Nombre Nave
            ws_res.cell(row=fila_res, column=4, value=f"={ref_hoja}!C6") #Tipo Nave
            
            trg = ws_res.cell(row=fila_res, column=5, value=f"={ref_hoja}!C8") #TRG
            trg.number_format = formato_entero
            trg.alignment = Alignment(horizontal="center", vertical="center")

            sitio= ws_res.cell(row=fila_res, column=6, value=f"={ref_hoja}!C9") #Sitio
            sitio.alignment = Alignment(horizontal="center", vertical="center")

            p_espia= ws_res.cell(row=fila_res, column=7, value=f"={ref_hoja}!C11") #Primera Espia
            p_espia.number_format = formato_fecha
            u_espia= ws_res.cell(row=fila_res, column=8, value=f"={ref_hoja}!C12") #Ultima Espia
            u_espia.number_format = formato_fecha
            # REVISAR
            t_oc= ws_res.cell(row=fila_res, column=9, value=f"={ref_hoja}!C14") #T.O.
            t_oc.number_format = formato_dec
            t_oc.alignment = Alignment(horizontal="center", vertical="center")

            b_dom= ws_res.cell(row=fila_res, column=10, value=f"={ref_hoja}!C18") #B.Dom
            b_dom.number_format = formato_dec
            b_dom.alignment = Alignment(horizontal="center", vertical="center")

            f_rend = ws_res.cell(row=fila_res, column=11, value=f"={ref_hoja}!C17") # F. Rend
            f_rend.number_format = formato_dec
            f_rend.alignment = Alignment(horizontal="center", vertical="center")

            tom= ws_res.cell(row=fila_res, column=12, value=f"={ref_hoja}!C19") #TOM
            tom.number_format = formato_dec
            tom.alignment = Alignment(horizontal="center", vertical="center")

            tnt= ws_res.cell(row=fila_res, column=13, value=f"={ref_hoja}!C20") #TNT
            tnt.number_format = formato_dec
            tnt.alignment = Alignment(horizontal="center", vertical="center")


            gap= ws_res.cell(row=fila_res, column=14, value=f"={ref_hoja}!C43") #GAP
            gap.number_format = formato_dec
            gap.alignment = Alignment(horizontal="center", vertical="center")

            multa= ws_res.cell(row=fila_res, column=15, value=f"={ref_hoja}!C45") #"Multa Inicial"
            multa.number_format = formato_dec
            multa.alignment = Alignment(horizontal="center", vertical="center")

            ton= ws_res.cell(row=fila_res, column=16, value=f"={ref_hoja}!C16") # Ton
            ton.number_format = formato_entero
            ton.alignment = Alignment(horizontal="center", vertical="center")

            vel = ws_res.cell(row=fila_res, column=17, value=f"={ref_hoja}!C47") # Vel
            vel.number_format = formato_dec
            vel.alignment = Alignment(horizontal="center", vertical="center")

            ws_res.cell(row=fila_res, column=18, value=f"={ref_hoja}!G16") #TIPO CARGA

            fila_res += 1
        ultima_fila_res = ws_res.max_row

        # Fila donde irá el total
        fila_total = ultima_fila_res + 2

        ws_res.cell(row=fila_total, column=4, value="TOTALES")  # Columna D
        ws_res.cell(row=fila_total, column=4).font = Font(name="Aptos Narrow", size=12, bold=True)
        ws_res.cell(row=fila_total, column=4).alignment = Alignment(horizontal="right", vertical="center")

        
        for celda in ws_res[fila_total][3:16]: # Accede directamente a la fila 47, columnas D(1) y P(2)
                celda.fill = relleno_gris
                celda.border = border_horizontal
                
        ws_res.cell(row=fila_total, column=4).border = Border(left=Side(border_style="thin", color="000000"),
                                                              top=Side(border_style="thin", color="000000"),
                                                              bottom=Side(border_style="thin", color="000000"))

        celda_total_multa = ws_res.cell(row=fila_total,column=15,value=f"=SUM(O2:O{ultima_fila_res})")
        celda_total_multa.border = border_completo
        celda_total_multa.number_format = formato_dec
        celda_total_multa.font = Font(name="Aptos Narrow", size=12, bold=True)
        celda_total_multa.alignment = Alignment(horizontal="center", vertical="center")

        celda_total_ton = ws_res.cell(row=fila_total, column=16, value=f"=SUM(P2:P{ultima_fila_res})")
        celda_total_ton.border = border_completo
        celda_total_ton.number_format = formato_entero
        celda_total_ton.font = Font(name="Aptos Narrow", size=12, bold=True)
        celda_total_ton.alignment = Alignment(horizontal="center", vertical="center")

        


        for fila in ws_res.iter_rows(
            min_row=2,
            max_row=ultima_fila_res,
            min_col=2,   # B
            max_col=18   # R
        ):
            for celda in fila:
                celda.border = border_completo

        # Guardar archivo
        ruta_final = obtener_ruta_salida(ruta_output_dir)
        wb.save(ruta_final)
        messagebox.showinfo("Éxito", f"Archivo creado en:\n{ruta_final}")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")
    
    return