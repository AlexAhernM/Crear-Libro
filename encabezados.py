
from openpyxl.styles import Font, Alignment, PatternFill


def crear_encabezados (ws_res, border_completo):

    # Encabezados hoja Res desde B1 hasta R1
    encabezados_res = [
        "Nº",
        "Nave",
        "Tipo Nave",
        "TRG",
        "Sitio",
        "Primera Espia",
        "Ultima Espia",
        "T.O",
        "B. Dom",
        "F.Rend",
        "TOM",
        "TNT",
        "GAP",
        "Multa TOM",
        "Ton",
        "T.E",
        "Multa TE",
        "Vel (ton/hr)",
        "TIPO DE CARGA"
    ]
    ws_res.column_dimensions['B'].width = 4
    ws_res.column_dimensions['C'].width = 24
    ws_res.column_dimensions['D'].width = 16
    ws_res.column_dimensions['E'].width = 10
    ws_res.column_dimensions['F'].width = 7
    ws_res.column_dimensions['G'].width = 16
    ws_res.column_dimensions['H'].width = 16
    ws_res.column_dimensions['I'].width = 8
    ws_res.column_dimensions['J'].width = 9
    ws_res.column_dimensions['K'].width = 9
    ws_res.column_dimensions['L'].width = 8
    ws_res.column_dimensions['M'].width = 8
    ws_res.column_dimensions['N'].width = 8
    ws_res.column_dimensions['O'].width = 12
    ws_res.column_dimensions['P'].width = 10
    ws_res.column_dimensions['Q'].width = 10
    ws_res.column_dimensions['R'].width = 10
    ws_res.column_dimensions['S'].width = 10
    ws_res.column_dimensions['T'].width = 35

    for i, texto in enumerate(encabezados_res, start=2):  # B=2 hasta T=20
        celda = ws_res.cell(row=1, column=i, value=texto)
        celda.font = Font(name="Aptos Narrow", size=10, bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        celda.border = border_completo

def obtener_columnas(df):
    return {
        "nave": df.columns[2],
        "servicio": df.columns[4],
        "tipo_nave": df.columns[5],
        "term": df.columns[6],
        "sitio": df.columns[7],
        "loa": df.columns[9],
        "trg": df.columns[11],
        "atraque": df.columns[12],
        "desatraque": df.columns[13],
        "ton": df.columns[31],
        "cu_met": df.columns[56],
        "zn": df.columns[57],
        "cu": df.columns[59],
        "cs": df.columns[61],
        "break": df.columns[64],
    }