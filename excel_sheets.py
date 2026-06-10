import pandas as pd
from openpyxl.styles import Alignment
from buscar_atras import buscar_datos_nave_hacia_atras
from encabezados import (formato_fecha, formato_dec, formato_entero, 
                         border_horizontal, border_completo, sin_bordes,  
                         relleno_rosa, relleno_celeste, relleno_gris, relleno_amarillo, 
                         fuente_liviana, alineacion_sangria, fuente_resultados, fuente_datos)

def crear_hojas(wb, df_filtrado, df,  col_nave, col_trg, col_loa, col_servicio, col_sitio, col_atraque, col_desatraque,
                col_tipo_nave, col_ton, col_cs, col_cu_met, col_break, col_zn, col_cu):
    hojas_naves = []

    for index, row in df_filtrado.iterrows():
                nombre_nave = str(row[col_nave]).strip()

                # Revisar si el TRG actual viene vacío
                trg_val = pd.to_numeric(row[col_trg], errors='coerce')
                loa_val = pd.to_numeric(row[col_loa], errors='coerce')

                datos_recuperados = None
                es_segunda_recalada_o_repetida = False

                if pd.isna(trg_val):
                    datos_recuperados = buscar_datos_nave_hacia_atras(
                        df=df,
                        index_actual=index,
                        col_nave=col_nave,
                        col_trg=col_trg,
                        col_loa=col_loa,
                        max_filas=20
                    )

                    if datos_recuperados is not None:
                        trg_val = datos_recuperados["trg"]
                        loa_val = datos_recuperados["loa"]
                        es_segunda_recalada_o_repetida = True

                # Crear nombre base de hoja
                sheet_name_base = "".join(x for x in nombre_nave if x.isalnum() or x in " -_")[:31]

                # Si se recuperó TRG desde una fila anterior, agregar (2), (3), etc.
                if es_segunda_recalada_o_repetida:
                    contador = 2

                    while True:
                        sufijo = f"({contador})"
                        largo_max_base = 31 - len(sufijo)
                        sheet_name = f"{sheet_name_base[:largo_max_base]}{sufijo}"

                        if sheet_name not in wb.sheetnames:
                            break

                        contador += 1
                else:
                    sheet_name = sheet_name_base

                    # Si por cualquier razón ya existe una hoja con ese nombre,
                    # también se numera para evitar error de Excel.
                    if sheet_name in wb.sheetnames:
                        contador = 2

                        while True:
                            sufijo = f"({contador})"
                            largo_max_base = 31 - len(sufijo)
                            sheet_name = f"{sheet_name_base[:largo_max_base]}{sufijo}"

                            if sheet_name not in wb.sheetnames:
                                break

                            contador += 1

                ws = wb.create_sheet(title=sheet_name)
                hojas_naves.append(sheet_name)

                # 1. Desactivar cuadrícula
                ws.sheet_view.showGridLines = False

                #aca deseo que desde la columna H hasta la columna G tengan un ancho de 10
                ws.column_dimensions['B'].width = 27
                ws.column_dimensions['C'].width = 21

                # Ajustar ancho de 10 para columnas desde la H hasta la O
                columnas_cuadro = ['G','H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
                for letra in columnas_cuadro:
                    ws.column_dimensions[letra].width = 10

                # --- Formato Numérico para el Cuadro de Bodegas (H20:O27) ---
                # Recorremos desde la fila 20 hasta la 27
                for r in range(20, 28): 
                    # Recorremos desde la columna 8 (H) hasta la 15 (O)
                    for c in range(7, 15): 
                        celda_cuadro = ws.cell(row=r, column=c)
                        
                        # Aplicamos el formato definido previamente
                        celda_cuadro.number_format = formato_dec
                        
                        # Opcional: Alinear a la derecha para que los decimales queden ordenados
                        celda_cuadro.alignment = Alignment(horizontal='right')

                
                # --- DATOS GENERALES (Texto) ---
                ws.cell(row=3, column=2, value="Nave")
                ws.cell(row=3, column=3, value=nombre_nave)

                ws.cell(row=4, column=2, value="Servicio")
                ws.cell(row=4, column=3, value=str(row[col_servicio]))

                ws.cell(row=5, column=2, value="Viaje")
                #ws.cell(row=5, column=3, value=row[col_viaje]) # Asumiendo columna de viaje

                ws.cell(row=17, column=4, value="Ton/hora")
                ws.cell(row=18, column=4, value="Ton")
                ws.cell(row=19, column=4, value="Hr")
                ws.cell(row=20, column=4, value="Hr")      

                c_loa = ws.cell(row=7, column=3, value=float(row[col_loa]))
                c_loa.number_format = formato_dec
                # LOA
                ws.cell(row=7, column=2, value="LOA")

                if pd.isna(loa_val):
                    c_loa = ws.cell(row=7, column=3, value="")
                else:
                    c_loa = ws.cell(row=7, column=3, value=float(loa_val))
                    c_loa.number_format = formato_dec

                # TRG
                ws.cell(row=8, column=2, value="TRG")

                if pd.isna(trg_val):
                    c_trg = ws.cell(row=8, column=3, value="")
                else:
                    c_trg = ws.cell(row=8, column=3, value=int(trg_val))
                    c_trg.number_format = formato_entero

                ws.cell(row=9, column=2, value="Sitio")        
                ws.cell(row=9, column=3, value=int(row[col_sitio]))

                ws.cell(row=10, column=2, value="Fecha Requerida de Atraque")
                c_fech_req = ws.cell(row=10, column=3)
                c_fech_req.number_format = formato_fecha

                ws.cell(row=11, column=2, value="Fecha amarra primera espia")        
                c_atraque = ws.cell(row=11, column=3, value=row[col_atraque])
                c_atraque.number_format = formato_fecha
                
                ws.cell(row=12, column=2, value="Fecha desamarra ultima espia")        
                c_desatraque = ws.cell(row=12, column=3, value=row[col_desatraque])
                c_desatraque.number_format = formato_fecha

                ws.cell(row=13, column=2, value="Tiempo de Espera")
                ws.cell(row=13, column=3, value='=IF(ISBLANK(C10),"",24*(C11-C10))')

                ws.cell(row=14, column=2, value="Tiempo de ocupación")
                tiempo_ocupacion = ws.cell(row=14, column=3, value="=24*(C12-C11)")
                tiempo_ocupacion.number_format = formato_dec

                for celda in ws["B14:C14"][0]: 
                    celda.fill = relleno_rosa

                ws.cell(row=16, column=2, value="Toneladas Transferidas")
                ton_val = pd.to_numeric(row[col_ton], errors="coerce")
                ton_val = float(0 if pd.isna(ton_val) else ton_val)

                tons = ws.cell(row=16, column=3, value=ton_val)
                tons.number_format = formato_entero
                
                ws.cell(row=17, column=2, value="Factor de Rendimiento")
                
                ws.cell(row=18, column=2, value="Bodega Dominante")
                ton_dominante = ws.cell(row=18, column=3, value="=MAX(G27:M27)")
                ton_dominante.number_format = formato_dec

                ws.cell(row=27, column=6, value="Totales")
            
            # --- DATOS GENERALES (Fila 6: Tipo Nave) ---
                ws.cell(row=6, column=2, value="Tipo Nave")
                
                # Paso 1: Convertir a número de forma segura (evita el error de base 10)
                tipo_nave_val = pd.to_numeric(row[col_tipo_nave], errors='coerce')
                #deseo que ceniza_soda sea numero decimal ¿como lo hago?
                tipo_nave_val = int(0 if pd.isna(tipo_nave_val) else tipo_nave_val) # Si es vacío, asume 0

                ceniza_soda = pd.to_numeric(row[col_cs], errors='coerce')
                ceniza_soda = float(0 if pd.isna(ceniza_soda) else ceniza_soda)
                if ceniza_soda>0:
                    ws.cell(row=16,column=7, value = "CENIZA DE SODA")
                    ws.cell(row=17, column=3, value=96)

                #porque siempre cobre_met=0, cuando en ocasiones no lo es
                #cobre metalico
                
                cobre_met = pd.to_numeric(row[col_cu_met], errors="coerce")
                cobre_met = int(0 if pd.isna(cobre_met) else cobre_met)
                print ("COBRE =", cobre_met, "toneladas", ton_val)
                
                #break bulk
                break_bulk = pd.to_numeric(row[col_break], errors="coerce")
                break_bulk = float(0 if pd.isna(break_bulk) else break_bulk)
                
                if break_bulk > 0 and abs(break_bulk - ton_val) < 0.01 and cobre_met==0:
                    ws.cell(row=16, column=7, value="BREAK BULK")
                    ws.cell(row=17, column=3, value=40)

                if break_bulk > 0 and abs(break_bulk - cobre_met) < 0.01:
                    ws.cell(row=16, column=7, value="COBRE METALICO")
                    ws.cell(row=17, column=3, value=120)

                conc_cu = pd.to_numeric(row[col_cu], errors="coerce")
                conc_cu = int(0 if pd.isna(conc_cu) else conc_cu)
                if conc_cu > 0 and conc_cu == ton_val:
                    ws.cell(row=16,column=7, value = "CONCENTRADO DE COBRE")
                    ws.cell(row=17, column=3, value=250)

                conc_zn = pd.to_numeric(row[col_zn], errors="coerce")
                conc_zn = int(0 if pd.isna(conc_zn) else conc_zn)
                if conc_zn > 0 and conc_zn == ton_val:
                    ws.cell(row=16,column=7, value = "CONCENtRADO DE ZINC")
                    ws.cell(row=17, column=3, value=250)

                if conc_zn > 0 and conc_cu >0:
                    ws.cell(row=16,column=7, value = "CONC. ZINC + CONC. CU")
                    ws.cell(row=17, column=3, value=250)

                # Caso Tipo Nave 32 (Container Ship)
                if tipo_nave_val == 32:
                    ws.cell(row=6, column=3, value="Container Ship")
                    ws.cell(row=16, column = 7, value = "CONTENEDORES")
                    # Sumamos columnas 32 a 39 (el límite 40 es para incluir la 39)
                    suma_embarque = pd.to_numeric(row.iloc[32:40], errors='coerce').fillna(0).sum()
                    ws.cell(row=17, column=3, value=21)
                    ws.cell(row=17, column=4, value="ctnr/hora")
                    ws.cell(row=18, column=4, value="ctnr")

                    teus=ws.cell(row=18, column=3, value=suma_embarque)
                    teus.number_format = formato_entero

                # Caso Tipo Nave 13 (Bulk Carrier)
                elif tipo_nave_val == 13:
                    ws.cell(row=6, column=3, value="Bulk Carrier")
                    
                    # Conversión segura para Zinc
                    val_zn = pd.to_numeric(row[col_zn], errors='coerce')
                    val_zn = 0 if pd.isna(val_zn) else val_zn
                    
                    if val_zn > 0:
                        ws.cell(row=17, column=3, value=250)

                elif tipo_nave_val == 11:
                    ws.cell(row=6, column=3, value="General Cargo Ship")

                ws.cell(row=19, column=2, value="Tiempo de Ocupacion Maximo")
                tom = ws.cell(row=19, column=3, value="=IFERROR(C18/C17, 0)+2")
                tom.number_format = formato_dec
                

                ws.cell(row=20, column=2, value="Tiempo de Inactividad")
                tnt = ws.cell(row=20, column=3, value="=SUM(C21:C27)") 
                tnt.number_format = formato_dec

                ws.cell(row=21, column=2, value="Atraque y Recepcion Nave")
                ws.cell(row=22, column=2, value="Colacion")
                ws.cell(row=23, column=2, value="Descanso Legal Amanecida")
                ws.cell(row=24, column=2, value="Tiempo Muerto Remate")
                ws.cell(row=25, column=2, value="Espera de carga / camiones")
                ws.cell(row=25, column=2, value="Draft")
                ws.cell(row=26, column=2, value="Inspeccion de Bodegas")
                ws.cell(row=27, column=2, value="")
                ws.cell(row=28, column=2, value="")
                ws.cell(row=29, column=2, value="")
                ws.cell(row=30, column=2, value="")
                ws.cell(row=31, column=2, value="")
                
                # Aplicar formato liviano + fondo amarillo al rango deseado
                # Cambia "B18:B31" por el rango que necesites resaltar
                for fila in ws["B21:C31"]:
                    for celda in fila:
                        celda.font = fuente_liviana
                        celda.alignment = alineacion_sangria
                        celda.fill = relleno_amarillo
                for fila in ws["C21:C31"]:
                    for celda in fila:
                        celda.number_format = formato_dec

                ws.cell(row=32, column=2, value="Tiempo Maximo Permitido (Real)")
                t_max= ws.cell(row=32, column=3, value="=C19+C20")
                t_max.number_format = '0.00'

                ws.cell(row=33, column=2, value="GAP")
                gap = ws.cell(row=33, column=3, value='=IF(C17=0,"Sin determinar aun",C32-C14)')
                gap.number_format = '0.00'

                ws.cell(row=34, column=2, value="*Si GAP (+) no aplica multa. Si GAP (-) aplica multa")

                for celda in ws["B34:C34"][0]: 
                    celda.fill = relleno_celeste

                tit_tom= ws.cell(row=36, column=2, value="Multa Tom")
                tit_tom.font = fuente_resultados

                ws.cell(row=37, column=2, value="Multa Base")
                ws.cell(row=37, column=3, value=0.035)

                ws.cell(row=38, column=2, value="Factor USPPI")
                ws.cell(row=38, column=3, value=1.803)

                ws.cell(row=39, column=2, value="Multa Final")
                m_final= ws.cell(row=39, column=3, value="=C37*C38")
                m_final.number_format = '0.000'
            
                calc_tom= ws.cell(row=41, column=2, value="Calculo Multa Tom")
                calc_tom.font = fuente_resultados

                ws.cell(row=42, column=2, value="Multa Final")
                ws.cell(row=42, column=3, value="=C39")

                ws.cell(row=43, column=2, value="Horas Exceso")
                ws.cell(row=43, column=3, value='=IFERROR(C33*-1, "Sin definir aun")')
                ws.cell(row=44, column=2, value="TRG")
                trg=ws.cell(row=44, column=3, value="=C8")
                trg.number_format = formato_entero

                ws.cell(row=45, column=2, value="MULTA TOM").border = border_horizontal
                m_apl = ws.cell(row=45, column=3, value='=IF(C20=0,"Sin Revision",IF(C43<0,0,C44*C43*C42))')
            
                m_apl.number_format = formato_dec
                m_apl.border = border_horizontal

                ws.cell(row=47, column=2, value="RENDIMIENTO").border = border_horizontal
                rend = ws.cell(row=47, column= 3, value = "=C16/(C14-C20)")
                rend.number_format = formato_dec
                rend.border = border_horizontal
            
                for celda in ws[45][1:3]: # Accede directamente a la fila 45, columnas B(1) y C(2)
                    celda.fill = relleno_rosa
                
                for celda in ws["B47:C47"][0]: 
                    celda.fill = relleno_celeste

                # --- CONFIGURACIÓN DEL CUADRO BODEGAS  (F18:N27) ---
                
                # 1. Definir los encabezados del cuadro (Fila 19, de F a N)
                encabezados_cuadro = ["", "BOD 1", "BOD 2", "BOD 3", "BOD 4", "BOD 5", "BOD 6", "BOD 7", "TOTALES"]
                

                for i, texto in enumerate(encabezados_cuadro):
                    celda = ws.cell(row=19, column=6 + i, value=texto)
                    celda.font = fuente_resultados
                    celda.alignment = Alignment(horizontal='right')
                    celda.border = border_horizontal

                # 4. Configuración del Cuerpo y Totales (Filas 19 a 27)
                for r in range(20, 29): # De 20 hasta 28
                    for c in range(6, 15): # Columnas F a N
                        celda_actual = ws.cell(row=r, column=c)
                        
                        if r == 27:
                            # Fila de TOTALES: Aplicar borde horizontal arriba y abajo
                            celda_actual.font = fuente_resultados
                            celda_actual.border = border_horizontal
                        else:
                            # Cuerpo de datos: Sin bordes, solo fuente normal
                            celda_actual.font = fuente_datos
                            celda_actual.border = sin_bordes

                #Revisar al parecer hay un error porque al abri excel envia un mensaje de error para reparar formula

                # 3. FÓRMULAS DE SUMA POR FILA (Columna N)
                # Sumar desde la columna F (6) hasta la M (13) para cada fila
                for r in range(19, 27): # Filas de datos (debajo del encabezado y sobre el total)
                    # Ejemplo: =SUM(F19:M19)
                    ws.cell(row=r, column=14).value = f"=SUM(G{r}:M{r})"

                # 4. FÓRMULAS DE SUMA POR COLUMNA (Fila 27)
                # Sumar verticalmente desde la fila 18 a la 26 para cada columna
                columnas_a_sumar = ['G', 'H', 'I', 'J', 'K', 'L', 'M','N']

                for letra in columnas_a_sumar:
                    # La 'G' es el índice 0 en nuestra lista. 
                    # Para que caiga en la columna 7 de Excel: 0 + 7 = 7.
                    col_num = columnas_a_sumar.index(letra) + 7                
                    # La suma debe ir desde la fila 20 (primer dato) hasta la 26 (último dato)
                    ws.cell(row=27, column=col_num).value =f"=SUM({letra}20:{letra}26)"

    return hojas_naves