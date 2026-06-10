from openpyxl.styles import Font, Alignment, Border, Side, PatternFill



    # ---------------formatos ---------

formato_fecha = 'dd/mm/yyyy h:mm'

formato_dec = '#,##0.00'

formato_entero = '#,##0'

# ----------------bordes -------------------

border_horizontal = Border(
top=Side(border_style="thin", color="000000"),
bottom=Side(border_style="thin", color="000000"))

border_completo = Border(
top=Side(border_style="thin", color="000000"),
bottom=Side(border_style="thin", color="000000"),
left=Side(border_style="thin", color="000000"),
right=Side(border_style="thin", color="000000")
)
sin_bordes = Border(
top=Side(border_style=None),
bottom=Side(border_style=None),
left=Side(border_style=None),
right=Side(border_style=None))

# ----------------rellenos --------------------

relleno_rosa =   PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

relleno_celeste = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")

relleno_gris =    PatternFill(start_color="F2F2F2", end_color = "F2F2F2", fill_type="solid") 

relleno_amarillo = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")                                                                                  

# ------------fuente y alineacion --------------
fuente_liviana = Font(name='Aptos Narrow', size=9, color='808080')
alineacion_sangria = Alignment(indent=1, vertical='center')

fuente_resultados = Font(name='Aptos NArrow', size=11, bold=True)
fuente_datos = Font(name = "Aptos Narrow", size = 10, bold = False)